# @InterpreterLocation : !C:\Users\20866\AppData\Local\Programs\Python\Python39\python.exe
# -*- coding: utf-8 -*-
# @Author : 20866
# @CreateTime : 2022/1/18 12:21:57
# @Project: Pycharm_Project_Set
# @File : send_email.py
# @Software : PyCharm
# @Site : https://gitee.com/qiongjulingjun

# 本文件对Excel进行相关的操作

import openpyxl
import random
import xlwt
import xlsxwriter


def main():
    # 创建一个工作簿
    wb = openpyxl.Workbook()

    # 创建一个表单
    wb.create_sheet('sheet_data')

    # 保存为xlsx类型的文件
    wb.save('text_1.xlsx')

    # 打开工作簿
    wb = openpyxl.load_workbook('../data_files/text_1.xlsx')

    # 选取表单
    sheet = wb['sheet_data']

    # 向里面填入数据

    # 制造数据：
    head_ = ['serial number', 'student ID', 'student name', 'Elective subjects']  # 表头
    num_ = []
    id_ = []
    name_ = []
    e_s_ = []
    list_all_ = [num_, id_, name_, e_s_]
    for i in range(1, 51):
        num_.append('00' + str(i))
        random_e_s = int(random.random() * 10)
        e_s_.append('course_' + str(random_e_s))
        if i < 10:
            id_.append('193009010' + str(i))
            name_.append('name_0' + str(i))
        else:
            id_.append('19300901' + str(i))
            name_.append('name_' + str(i))
    # 填数据：
    s = 0
    head_list = ['A', 'B', 'C', 'D']
    for items in head_list:
        sheet[items + '1'] = head_[s]
        s += 1

    for i in range(2, 52):
        sheet['A' + str(i)] = num_[i - 2]

    for j in range(2, 52):
        sheet['B' + str(j)] = id_[j - 2]

    for k in range(2, 52):
        sheet['C' + str(k)] = name_[k - 2]

    for m in range(2, 52):
        sheet['D' + str(m)] = e_s_[m - 2]
    wb.save('text_1.xlsx')
    wb.close()


def write_xlwt():
    # 创建工作簿：
    my_workbook = xlwt.Workbook()

    # 添加表
    my_sheet = my_workbook.add_sheet('test_1')

    # 写数据
    my_sheet.write(0, 0, 'Hello')
    my_sheet.write(0, 1, 'World')
    my_sheet.write(0, 2, 'How')
    my_sheet.write(0, 3, 'Are')
    my_sheet.write(1, 0, '小卜')
    my_sheet.write(1, 1, '小牛')
    my_workbook.save('test_1_1.xls')
    pass


def xlsx_writer():
    # 写Excel
    my_workbook = xlsxwriter.Workbook('../data_files/test_2_1.xlsx')
    my_sheet = my_workbook.add_worksheet('text_1')
    my_sheet.set_column('A:A', 20)  # 设置第一列宽度为20
    my_sheet.set_column('B:B', 20)  # 设置第二列宽度为20

    bold = my_workbook.add_format({'bold': True})  # 设置一个加粗的对象
    my_sheet.write('A1', 'Num')  # 在A1单元格上写入 num
    my_sheet.write('B1', 'ID', bold)  # 填入元素并且加粗
    my_sheet.write('A2', '中文加粗', bold)
    my_sheet.write('B2', '中文加粗2', bold)
    my_sheet.write(4, 0, 65)
    my_sheet.write(4, 1, 66)
    my_sheet.write(4, 2, '=SUM(A5:B5)')
    my_workbook.close()
    pass


if __name__ == '__main__':
    # main()
    # write_xlwt()
    xlsx_writer()
    pass

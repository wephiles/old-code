# -*- coding: utf-8 -*-
# @CreateTime : 2024/4/15 015 9:05
# @Author : 瑾瑜@20866
# @IDE : PyCharm
# @File : studyProject/excel_demo3.py
# @Description : 
# @Interpreter : python 3.10
# @Motto : You must take your place in the circle of life!
# @Site : https://github.com/wephiles or https://gitee.com/wephiles

# import os
# from pprint import pprint
# from openpyxl import load_workbook
#
# FILE_FOLDER = os.path.join("files", "营收报表", "2020年1月份营收.xlsx")
#
# # 获取workbook对象
# workbook_obj = load_workbook(FILE_FOLDER)
#
# sheet_obj = workbook_obj.worksheets[0]
#
# cell_obj = sheet_obj.cell(row=1, column=1)
# print(cell_obj.value)
#
# cell_obj.value = "age"
# print(cell_obj.value)
#
# # 将内存中的数据写入到磁盘
# workbook_obj.save(os.path.join("files",  "2020年1月份营收-修改.xlsx"))

# # 新建excel文件
#
# import os
# from openpyxl.workbook import Workbook
#
# FILE_FOLDER = os.path.join("files", "营收报表", "2020年1月份营收.xlsx")
# workbook_obj = Workbook()
# # sheet_obj = workbook_obj.worksheets[0]
# # cell_obj = sheet_obj.cell(row=1, column=1)
# #
# # cell_obj.value = "age"
# workbook_obj.save("./files/create_1.xlsx")

# import os
# from openpyxl.workbook import Workbook
#
# FILE_FOLDER = os.path.join("files", "营收报表", "2020年1月份营收.xlsx")
#
# # 创建一个workbook对象 创建好以后会自动生成一个默认的worksheet--sheet1
# workbook_obj = Workbook()
# sheet_0 = workbook_obj.worksheets[0]
# sheet_0.title = "数据集"
# sheet_0.cell(row=1, column=1, value="A")
# sheet_0.cell(row=1, column=2, value="B")
# sheet_0.cell(row=1, column=3, value="C")
# sheet_0.cell(row=1, column=4, value="D")
#
# # 创建sheet
# sheet_1 = workbook_obj.create_sheet(title="上海", index=1)
# sheet_1.cell(row=1, column=1, value="A")
# sheet_1.cell(row=1, column=2).value = "B"
# sheet_1.cell(row=1, column=3).value = "C"
# sheet_1.cell(row=1, column=4).value = "D"
#
# sheet_2 = workbook_obj.create_sheet(title="北京", index=2)
# sheet_3 = workbook_obj.create_sheet(title="广州", index=3)
# sheet_4 = workbook_obj.create_sheet(title="深圳", index=4)
#
# workbook_obj.save("./files/create_1.xlsx")

# import os
# from openpyxl.workbook import Workbook
#
# FILE_FOLDER = os.path.join("files", "营收报表", "2020年1月份营收.xlsx")
#
# # 创建一个workbook对象 创建好以后会自动生成一个默认的worksheet
# workbook_obj = Workbook()
# sheet_0 = workbook_obj.worksheets[0]
#
# new_sheet = workbook_obj.copy_worksheet(sheet_0)
#
# new_sheet.title = "备份"
#
# workbook_obj.save("./files/create_new.xlsx")


# import os
# from openpyxl import load_workbook
#
# file_name = os.path.join("files", "create_new.xlsx")
# # 1. 获取对象
# workbook_obj = load_workbook(file_name)
#
# del workbook_obj["备份"]
#
# workbook_obj.save("./files/create_new.xlsx")

# from openpyxl.styles import Alignment
# from openpyxl.workbook import Workbook
#
# workbook_obj = Workbook()
#
# sheet_0 = workbook_obj.worksheets[0]
#
# # 居中 换行
# cell_obj = sheet_0.cell(1, 1)
# cell_obj.value = "学生ID学生ID学生ID学生ID学生ID"
# cell_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 设置居中、wrap_text自动换行
#
# workbook_obj.save("./files/op_cell.xlsx")


from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, Color
from openpyxl.workbook import Workbook

workbook_obj = Workbook()

sheet_0 = workbook_obj.worksheets[0]

name_list = ["张三", "里斯", "王五", "二麻子"]

# 设置行高 宽度
sheet_0.row_dimensions[1].height = 50
sheet_0.column_dimensions["A"].height = 80
sheet_0.column_dimensions["B"].height = 100

# 居中 换行
sheet_0.cell(row=1, column=1).value = "姓名"
sheet_0.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
sheet_0.cell(row=1, column=1).fill = PatternFill(patternType='solid', fgColor=Color(rgb="808080"))

for row, text in enumerate(name_list, 2):
    cell_obj = sheet_0.cell(row, 1)
    cell_obj.value = text
    cell_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 设置居中、wrap_text自动换行
    cell_obj.border = Border(top=Side(style="thin", color="0000FF"),
                             left=Side(style="thin", color="0000FF"),
                             right=Side(style="medium", color="0000FF"),
                             bottom=Side(style="medium", color="0000FF"))
    cell_obj.font = Font(name="宋体", size=12, bold=True, italic=False, color="FFFF00")

workbook_obj.save("./files/op_cell.xlsx")

# from openpyxl.styles import Alignment, Border, Side
# from openpyxl.workbook import Workbook
#
# workbook_obj = Workbook()
#
# sheet_0 = workbook_obj.worksheets[0]
#
# # 操作cell方式一
# cell_obj = sheet_0.cell(1, 1)
# cell_obj.value = "学生ID"
#
# cell_obj.border = Border(top=Side(style="thin", color="0000FF"),
#                          left=Side(style="thin", color="0000FF"),
#                          right=Side(style="thin", color="0000FF"),
#                          bottom=Side(style="thin", color="0000FF"))
#
# workbook_obj.save("./files/op_cell.xlsx")

# --END--

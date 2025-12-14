# -*- coding: utf-8 -*-
# @CreateTime : 2024/4/14 014 20:34
# @Author : 瑾瑜@20866
# @IDE : PyCharm
# @File : studyProject/excel_practice.py
# @Description : 
# @Interpreter : python 3.10
# @Motto : You must take your place in the circle of life!
# @Site : https://github.com/wephiles or https://gitee.com/wephiles

# import os
# import re
# from openpyxl import load_workbook
#
# file_path = os.path.join("files", "Amtrak.xlsx")
#
# workbook_obj = load_workbook(file_path)
#
# sheet_obj = workbook_obj.worksheets[0]
#
# for row_obj in sheet_obj.iter_rows(min_row=2):
#     time_str = row_obj[6].value
#     if not time_str:
#         continue
#     time_num = int(re.findall(r"(\d+)分钟", time_str)[0])
#     if time_num > 10:
#         print(time_num)


# from openpyxl import load_workbook
#
#
# wb = load_workbook('./files/bankloan.xlsx')
# sheet_obj = None
# sum_ = 0
# sheet_names = wb.sheetnames
# for i in range(0, len(sheet_names)):
#     if sheet_names[i] == "bankloan":
#         sheet_obj = wb.worksheets[i]
#
# for row_obj in sheet_obj.iter_rows(min_row=2):
#     cell_income_obj = row_obj[4]
#     if not cell_income_obj.value:
#         continue
#     sum_ += cell_income_obj.value
#     print(cell_income_obj.value)
# print("sum:", sum_)

# import os
# import re
# from openpyxl import load_workbook
#
# file_path = os.path.join("files", "Amtrak.xlsx")
#
# workbook_obj = load_workbook(file_path)
#
# result = {}
#
# for sheet_obj in workbook_obj.worksheets:
#     sheet_name = sheet_obj.title
#
#     for row_list in sheet_obj.iter_rows(min_row=2):
#         count = row_list[1].value
#         if not count:
#             continue
#         total_count += count
#     result[sheet_name] = total_count
# print(result)

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Color, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


workbook_obj = Workbook()

sheet_obj = workbook_obj.worksheets[0]
sheet_obj.title = "用户信息"

header_name_list = ["用户名", "密码", "个人邮箱"]

for i in range(1, 4):
    sheet_obj.column_dimensions[get_column_letter(i)].width = 30

sheet_obj.cell(row=1, column=1).value = "瑾瑜"
sheet_obj.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
sheet_obj.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

# 表头 居中 字体 背景颜色 边框
for col, header_name in enumerate(header_name_list, start=1):
    sheet_obj.cell(row=2, column=col).value = header_name
    sheet_obj.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
    sheet_obj.cell(row=2, column=col).fill = PatternFill(fgColor="DCDCDC", patternType="solid")
    sheet_obj.cell(row=2, column=col).font = Font(name="楷体", bold=True, size=12, color=Color(rgb="FF0000"))
    sheet_obj.cell(row=2, column=col).border = Border(left=Side(border_style="medium", color="000000"),
                                                      right=Side(border_style="medium", color="000000"),
                                                      top=Side(border_style="medium", color="000000"),
                                                      bottom=Side(border_style="medium", color="000000"),
                                                      )

with open("./db/users/account.txt", "r", encoding='utf-8') as fp:
    line_num = 3
    for line in fp:
        line = line.strip()
        if not line:
            continue
        user_msg = line.split(",")
        for col, msg in enumerate(user_msg, start=1):
            sheet_obj.cell(line_num, col).value = msg
        line_num += 1

workbook_obj.save("./files/db/user_info.xlsx")

# --END--

# -*- coding: utf-8 -*-
# @CreateTime : 2024/4/16 016 14:21
# @Author : 瑾瑜@20866
# @IDE : PyCharm
# @File : studyProject/excel_append.py
# @Description : 
# @Interpreter : python 3.10
# @Motto : You must take your place in the circle of life!
# @Site : https://github.com/wephiles or https://gitee.com/wephiles

import os
from openpyxl import load_workbook, Workbook


def write_excel(user_name, password):
    """
    将用户名和密码写入excel文件。
    Args:
        user_name (str): 用户名
        password (str): 密码

    Returns:

    """
    # 1. 文件是否存在
    file_path = os.path.join("files", "db.xlsx")
    if os.path.exists(file_path):
        # 2. 打开文件
        workbook_obj = load_workbook(file_path)
    else:
        # 3. 创建文件
        workbook_obj = Workbook()
    sheet_obj = workbook_obj.worksheets[0]

    # 4. 写入数据
    first_row_data = sheet_obj.cell(row=1, column=1).value
    if not first_row_data:  # 第一行没有数据
        row_index = 1
    else:  # 第一行有数据
        row_index = sheet_obj.max_row + 1

    sheet_obj.cell(row=row_index, column=1).value = user_name
    sheet_obj.cell(row=row_index, column=2).value = password
    workbook_obj.save(file_path)


def run():
    """
    程序入口。
    Returns:

    """
    while True:
        print("输入用户名和密码，输入q/Q退出:")
        user_name = input("用户名 >>> ")
        password = input("密  码 >>> ")
        if user_name == "q" or user_name == "q":
            break
        write_excel(user_name, password)


if __name__ == '__main__':
    run()

# --END--

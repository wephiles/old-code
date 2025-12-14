# -*- coding: utf-8 -*-
# @CreateTime : 2024/4/14 014 18:17
# @Author : 瑾瑜@20866
# @IDE : PyCharm
# @File : studyProject/excel_demo1.py
# @Description : 
# @Interpreter : python 3.10
# @Motto : You must take your place in the circle of life!
# @Site : https://github.com/wephiles or https://gitee.com/wephiles

# import os
# from openpyxl import load_workbook
#
# file_name = os.path.join("files", "Amtrak.xlsx")
# # 1. 获取对象
# workbook_obj = load_workbook(file_name)
#
# # 2. 读取workbook里面的所有sheet名字
# sheet_list= workbook_obj.sheetnames
# print(sheet_list)
#
# v2 = workbook_obj.worksheets
# print(v2)

# import os
# from openpyxl import load_workbook
#
# file_name = os.path.join("files", "Amtrak.xlsx")
# # 1. 获取对象
# workbook_obj = load_workbook(file_name)
#
# # 2. 读取workbook里面的所有sheet对象
# sheet_obj_list = workbook_obj.worksheets
#
# # 3. 获取某一个sheet对象
# sheet_obj = sheet_obj_list[0]
#
# # 4. 读取sheet中的单元格对象 -- 方式1
# cell_obj = sheet_obj.cell(row=1, column=1)
#
# # 4. 读取sheet中的单元格对象 -- 方式2
# cell_obj_1 = sheet_obj["A1"]
#
# # # 5. 读取cell对象中的文本
# # print(cell_obj.value)
# # print(cell_obj_1.value)
#
# # # 6. 读取一行
# # row_tuple = sheet_obj[1]
# # print(row_tuple)
# #
# # for cell_obj in row_tuple:
# #     print(cell_obj.value)
#
# # # 7. 读取所有行
# #
# # for row in sheet_obj.rows:
# #     row_text_list = []
# #     for cell_obj in row:
# #         row_text_list.append(cell_obj.value)
# #     print(row_text_list)
#
# # 8. 读取所有列
# for row in sheet_obj.rows:
#     cell_obj_0 = row[0]
#     cell_obj_1 = row[1]
#     print(cell_obj_0.value, cell_obj_1.value)


# import os
# from openpyxl import load_workbook
# from openpyxl.cell.cell import Cell, MergedCell
#
# file_name = os.path.join("files", "Amtrak.xlsx")
# # 1. 获取对象
# workbook_obj = load_workbook(file_name)
#
# # 2. 读取workbook里面的所有sheet对象
# sheet_obj_list = workbook_obj.worksheets
#
# # 3. 获取某一个sheet对象
# sheet_obj = sheet_obj_list[3]
#
# # # 10. 获取合并的单元格
# # cell_1_obj = sheet_obj.cell(1, 1)
# # cell_2_obj = sheet_obj.cell(1, 2)
# #
# # print(cell_1_obj, cell_2_obj)
# # print(cell_1_obj.value, cell_2_obj.value)
#
# # # 原始内容:
# # for row in sheet_obj.rows:
# #     text_list = []
# #     for cell in row:
# #         text_list.append(cell.value)
# #     print(text_list)
#
# # # 如果是被合并的单元格 让其默认值等于 -
# # for row in sheet_obj.rows:
# #     text_list = []
# #     for cell in row:
# #         # 判断 这个单元格是Cell对象还是MergeCell对象
# #         if type(cell) is MergedCell:
# #             text_list.append("-")
# #         elif type(cell) is Cell:
# #             text_list.append(cell.value)
# #         else:
# #             text_list.append("Dont Know")
# #     print(text_list)
#
#
# # 获取当前sheet中所有的合并的单元格
#
# def get_merged_cell(coordinate, sheet_obj):
#     """
#
#     Args:
#         coordinate (): 坐标 需要这种形式: "A5"
#         sheet_obj (): 某一个sheet对象
#
#     Returns:
#
#     """
#     for item in sheet_obj.merged_cells:
#         # print(item, type(item))
#         if coordinate in item:
#             return item.start_cell.value
#
#
# # 如果是被合并的单元格 让其默认值等于第一个单元格的值
# for row in sheet_obj.rows:
#     text_list = []
#     for cell in row:
#         # 判断 这个单元格是Cell对象还是MergeCell对象
#         if type(cell) is MergedCell:
#             # 获取单元格的内容:
#             merged_text = get_merged_cell(cell.coordinate, sheet_obj)
#             text_list.append(merged_text)
#         elif type(cell) is Cell:
#             text_list.append(cell.value)
#         else:
#             text_list.append("Dont Know")
#     print(text_list)

# import os
# from openpyxl import load_workbook
# from openpyxl.cell.cell import Cell, MergedCell
#
# file_name = os.path.join("files", "Amtrak.xlsx")
# # 1. 获取对象
# workbook_obj = load_workbook(file_name)
#
# # 2. 读取workbook里面的所有sheet对象
# sheet_obj_list = workbook_obj.worksheets
#
# # 3. 获取某一个sheet对象
# sheet_obj = sheet_obj_list[3]
#
# # 从第n行开始读 读到第几行
# for row in sheet_obj.iter_rows(min_row=3, max_row=4):
#     print(row[0].value)

import os
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

file_name = os.path.join("files", "Amtrak.xlsx")
# 1. 获取对象
workbook_obj = load_workbook(file_name)

# 2. 读取workbook里面的所有sheet对象
sheet_obj_list = workbook_obj.worksheets

# 3. 获取某一个sheet对象
sheet_obj = sheet_obj_list[3]

# 总共有几行 总共有几列
print(sheet_obj.max_row)
print(sheet_obj.max_column)

# --END--

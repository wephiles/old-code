# -*- coding: utf-8 -*-
# @CreateTime : 2024/4/15 015 21:06
# @Author : 瑾瑜@20866
# @IDE : PyCharm
# @File : studyProject/merge_sheet.py
# @Description : 
# @Interpreter : python 3.10
# @Motto : You must take your place in the circle of life!
# @Site : https://github.com/wephiles or https://gitee.com/wephiles

from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment

workbook_obj = Workbook()
sheet_obj = workbook_obj.worksheets[0]
cell_obj = sheet_obj.cell(row=1, column=1)
cell_obj.value = "中华人民共和国万岁"
cell_obj.alignment = Alignment(horizontal="center", vertical="center")

# Merge cells A1:C1
sheet_obj.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

workbook_obj.save("./files/merge_sheet.xlsx")

# --END--

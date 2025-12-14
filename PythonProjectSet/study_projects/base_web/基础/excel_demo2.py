# -*- coding: utf-8 -*-
# @CreateTime : 2024/4/14 014 21:57
# @Author : 瑾瑜@20866
# @IDE : PyCharm
# @File : studyProject/excel_demo2.py
# @Description : 
# @Interpreter : python 3.10
# @Motto : You must take your place in the circle of life!
# @Site : https://github.com/wephiles or https://gitee.com/wephiles

# import os
# from pprint import pprint
# from openpyxl import load_workbook
#
# # 列出所有的文件名
#
# result = {}
#
#
# file_folder = os.path.join("files", "营收报表")
# file_list = os.listdir(file_folder)
#
# # 遍历所有文件
# for file_name_str in file_list:
#     if not file_name_str.startswith("2020"):
#         continue
#     # 去掉后缀名，形成字典的key
#     file_delete_postfix_name = file_name_str.split(".")[0]
#
#     # 初始化result字典
#     result[file_delete_postfix_name] = dict()
#
#     # 初始化工作簿对象
#     workbook_obj = load_workbook(os.path.join(file_folder, file_name_str))
#
#     # 遍历一个文件里面的所有sheet
#     for sheet_name in workbook_obj.sheetnames:
#         # workbook_obj.sheetnames ： 是个列表，里面包含了所有的sheet名称，每个元素都是字符串
#         sum_ = 0  # 用以计算总收入
#
#         # 获取sheet对象
#         sheet_obj = workbook_obj[sheet_name]
#         # 遍历所有行
#         for row_obj in sheet_obj.iter_rows(min_row=2):
#             num_income = row_obj[4].value  # 获取每一行的数据
#             if not num_income:
#                 continue
#             sum_ += int(num_income)  # 不空的话求和
#         result[file_delete_postfix_name][sheet_name] = sum_  # 保存数据
# pprint(result)


import os
from pprint import pprint
from openpyxl import load_workbook

FILE_FOLDER = os.path.join("files", "营收报表")


def run():
    result = {}

    # file_folder = os.path.join("files", "营收报表")
    file_list = os.listdir(FILE_FOLDER)
    # 遍历所有文件
    for file_name_str in file_list:
        if not file_name_str.startswith("2020"):
            continue
        # 去掉后缀名，形成字典的key
        file_delete_postfix_name = file_name_str.split(".")[0]

        # 初始化result字典
        result[file_delete_postfix_name] = dict()

        # 初始化工作簿对象
        workbook_obj = load_workbook(os.path.join(FILE_FOLDER, file_name_str))

        # 遍历一个文件里面的所有sheet
        for sheet_name in workbook_obj.sheetnames:
            # workbook_obj.sheetnames ： 是个列表，里面包含了所有的sheet名称，每个元素都是字符串
            sum_ = 0  # 用以计算总收入

            # 获取sheet对象
            sheet_obj = workbook_obj[sheet_name]
            # 遍历所有行
            for row_obj in sheet_obj.iter_rows(min_row=2):
                num_income = row_obj[4].value  # 获取每一行的数据
                if not num_income:
                    continue
                sum_ += int(num_income)  # 不空的话求和
            result[file_delete_postfix_name][sheet_name] = sum_  # 保存数据
    pprint(result)


if __name__ == '__main__':
    run()

# --END--
